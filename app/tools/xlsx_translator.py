from __future__ import annotations

import copy
import io
import hashlib
import json
import posixpath
import re
import sqlite3
import time
import unicodedata
import urllib.parse
import uuid
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

from app.settings import PROCESSED_DIR, STORAGE_DIR


from app.constants import (
    DEFAULT_MODELS,
    LANGUAGES,
    PROVIDERS,
    DEFAULT_MAX_WORKERS,
    DEFAULT_MAX_ITEMS,
    DEFAULT_MAX_CHARS,
)


VIETNAMESE_DIACRITIC_RE = re.compile(
    r"[àáạảãâầấậẩẫăằắặẳẵèéẹẻẽêềếệểễìíịỉĩòóọỏõôồốộổỗơờớợởỡ"
    r"ùúụủũưừứựửữỳýỵỷỹđ]",
    re.IGNORECASE,
)

SCRIPT_RANGES = {
    "Japanese": re.compile(r"[\u3040-\u30ff\u3400-\u9fff]"),
    "Korean": re.compile(r"[\uac00-\ud7af]"),
    "Chinese Simplified": re.compile(r"[\u3400-\u9fff]"),
    "Thai": re.compile(r"[\u0e00-\u0e7f]"),
}

TARGET_LANGUAGE_MARKERS = {
    "Vietnamese": {
        "cua",
        "cho",
        "cac",
        "khong",
        "duoc",
        "trong",
        "voi",
        "hang",
        "san",
        "pham",
        "mo",
        "ta",
        "ten",
        "gia",
        "ban",
        "ngay",
        "thang",
    },
    "English": {
        "the",
        "and",
        "for",
        "with",
        "from",
        "this",
        "that",
        "product",
        "description",
        "name",
        "price",
        "size",
        "color",
    },
    "French": {
        "le",
        "la",
        "les",
        "des",
        "une",
        "pour",
        "avec",
        "dans",
        "produit",
        "description",
    },
    "Spanish": {
        "el",
        "la",
        "los",
        "las",
        "para",
        "con",
        "producto",
        "descripcion",
        "precio",
    },
}

XML_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"

ET.register_namespace("", XML_NS)
ET.register_namespace("r", REL_NS)


class TranslationError(RuntimeError):
    pass


class XlsxXmlFallbackRequired(RuntimeError):
    pass


class ProviderRequestError(TranslationError):
    def __init__(self, status_code: int, detail: str):
        super().__init__(f"Provider request failed ({status_code}): {detail[:800]}")
        self.status_code = status_code
        self.detail = detail
        self.provider_code = extract_provider_error_code(detail)


def list_sheet_names(xlsx_bytes: bytes) -> list[str]:
    try:
        with zipfile.ZipFile(io.BytesIO(xlsx_bytes), "r") as archive:
            workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
            sheets_node = workbook_root.find(qn("sheets"))
            if sheets_node is None:
                raise XlsxXmlFallbackRequired("workbook has no sheets node")
            names = [sheet.get("name", "") for sheet in sheets_node.findall(qn("sheet"))]
            return [name for name in names if name]
    except Exception:
        pass

    workbook = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=False)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def translate_workbook(
    *,
    xlsx_bytes: bytes,
    original_name: str,
    selected_sheets: list[str],
    target_language: str,
    provider: str,
    model: str,
    api_key: str,
    base_url: str = "",
    max_workers: int = DEFAULT_MAX_WORKERS,
    max_items: int = DEFAULT_MAX_ITEMS,
) -> tuple[str, Path, list[str]]:
    logs: list[str] = []
    try:
        return translate_workbook_xml(
            xlsx_bytes=xlsx_bytes,
            original_name=original_name,
            selected_sheets=selected_sheets,
            target_language=target_language,
            provider=provider,
            model=model,
            api_key=api_key,
            base_url=base_url,
            max_workers=max_workers,
            max_items=max_items,
            logs=logs,
        )
    except XlsxXmlFallbackRequired as exc:
        logs.append(f"XML fast path fallback: {exc}")
        return translate_workbook_openpyxl(
            xlsx_bytes=xlsx_bytes,
            original_name=original_name,
            selected_sheets=selected_sheets,
            target_language=target_language,
            provider=provider,
            model=model,
            api_key=api_key,
            base_url=base_url,
            max_workers=max_workers,
            max_items=max_items,
            prefix_logs=logs,
        )


def translate_workbook_openpyxl(
    *,
    xlsx_bytes: bytes,
    original_name: str,
    selected_sheets: list[str],
    target_language: str,
    provider: str,
    model: str,
    api_key: str,
    base_url: str = "",
    max_workers: int = DEFAULT_MAX_WORKERS,
    max_items: int = DEFAULT_MAX_ITEMS,
    prefix_logs: list[str] | None = None,
) -> tuple[str, Path, list[str]]:
    if provider not in PROVIDERS:
        raise TranslationError(f"Unsupported provider: {provider}")
    if not api_key:
        raise TranslationError("Missing API key. Enter a key or configure the default key on the server.")

    system_start_time = time.time()
    
    logs = list(prefix_logs or [])
    logs.extend([
        "Engine: openpyxl fallback",
        f"Provider: {PROVIDERS[provider]}",
        f"Model: {model}",
        f"Target language: {target_language}",
    ])
    workbook = load_workbook(io.BytesIO(xlsx_bytes), data_only=False)
    try:
        sheet_names = set(workbook.sheetnames)
        targets = selected_sheets or list(workbook.sheetnames)
        missing = [name for name in targets if name not in sheet_names]
        if missing:
            raise TranslationError(f"Sheet not found: {', '.join(missing)}")

        cells: list[tuple[Any, str, str]] = []
        scanned_cells = 0
        candidate_cells = 0
        skipped_target_language = 0
        for sheet_name in targets:
            sheet = workbook[sheet_name]
            count_before = len(cells)
            skipped_before = skipped_target_language
            for row in sheet.iter_rows():
                for cell in row:
                    scanned_cells += 1
                    value = cell.value
                    if should_translate(value, cell.data_type):
                        candidate_cells += 1
                        original_text = value
                        normalized_text = normalize_text(value)
                        if is_likely_target_language(normalized_text, target_language):
                            skipped_target_language += 1
                            continue
                        cells.append((cell, original_text, normalized_text))
            logs.append(
                f"Sheet '{sheet_name}': queued {len(cells) - count_before} text cell(s); "
                f"skipped {skipped_target_language - skipped_before} likely already in {target_language}."
            )

        if not cells and not skipped_target_language:
            raise TranslationError("No translatable text cells found in the selected sheet(s).")

        unique_texts = list(dict.fromkeys(text for _, _, text in cells))
        duplicate_cells = len(cells) - len(unique_texts)
        logs.append(f"Scanned {scanned_cells} workbook cell(s).")
        logs.append(f"Found {candidate_cells} text cell(s) after base filters.")
        if skipped_target_language:
            logs.append(f"Skipped {skipped_target_language} cell(s) likely already in {target_language}.")
        logs.append(
            f"Deduplicated {len(cells)} queued cell(s) to {len(unique_texts)} unique text(s); "
            f"saved {duplicate_cells} duplicate translation(s)."
        )

        if not cells:
            job_id, output_path = save_workbook(workbook, original_name, target_language)
            logs.append("No cells were sent to the provider because all candidates looked already translated.")
            logs.append(f"Saved workbook without translation changes: {output_path.name}")
            system_elapsed = time.time() - system_start_time
            logs.append(f"✅ Total system execution time: {system_elapsed:.2f} seconds.")
            return job_id, output_path, logs

        translated, cache_hits, cache_misses = translate_texts_with_cache(
            texts=unique_texts,
            target_language=target_language,
            provider=provider,
            model=model,
            api_key=api_key,
            base_url=base_url,
            logs=logs,
            max_workers=max_workers,
            max_items=max_items,
        )
        logs.append(f"Translation cache: {cache_hits} hit(s), {cache_misses} miss(es).")
        translations_by_text = dict(zip(unique_texts, translated))

        for cell, original_text, normalized_text in cells:
            cell.value = preserve_wrapping(original_text, translations_by_text[normalized_text])

        job_id, output_path = save_workbook(workbook, original_name, target_language)
        logs.append(f"Saved translated workbook: {output_path.name}")
        
        system_elapsed = time.time() - system_start_time
        logs.append(f"✅ Total system execution time: {system_elapsed:.2f} seconds.")
        
        return job_id, output_path, logs
    finally:
        workbook.close()


def translate_workbook_xml(
    *,
    xlsx_bytes: bytes,
    original_name: str,
    selected_sheets: list[str],
    target_language: str,
    provider: str,
    model: str,
    api_key: str,
    base_url: str = "",
    max_workers: int = DEFAULT_MAX_WORKERS,
    max_items: int = DEFAULT_MAX_ITEMS,
    logs: list[str] | None = None,
) -> tuple[str, Path, list[str]]:
    if provider not in PROVIDERS:
        raise TranslationError(f"Unsupported provider: {provider}")
    if not api_key:
        raise TranslationError("Missing API key. Enter a key or configure the default key on the server.")

    logs = logs if logs is not None else []
    system_start_time = time.time()
    logs.extend(
        [
            "Engine: XLSX XML fast path",
            f"Provider: {PROVIDERS[provider]}",
            f"Model: {model}",
            f"Target language: {target_language}",
        ]
    )

    try:
        archive = zipfile.ZipFile(io.BytesIO(xlsx_bytes), "r")
    except zipfile.BadZipFile as exc:
        raise XlsxXmlFallbackRequired("uploaded file is not a readable XLSX zip") from exc

    with archive:
        names = set(archive.namelist())
        if "xl/workbook.xml" not in names or "xl/_rels/workbook.xml.rels" not in names:
            raise XlsxXmlFallbackRequired("workbook relationships are missing")

        phase_start = time.time()
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        sheet_paths = resolve_sheet_paths(workbook_root, rels_root)
        targets = selected_sheets or list(sheet_paths)
        missing = [name for name in targets if name not in sheet_paths]
        if missing:
            raise TranslationError(f"Sheet not found: {', '.join(missing)}")

        target_paths = [sheet_paths[name] for name in targets]
        missing_paths = [path for path in target_paths if path not in names]
        if missing_paths:
            raise XlsxXmlFallbackRequired(f"worksheet XML is missing: {', '.join(missing_paths)}")
        shared_root = None
        shared_items: list[ET.Element] = []
        shared_text_by_index: dict[int, str] = {}
        if "xl/sharedStrings.xml" in names:
            shared_root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            shared_items = list(shared_root.findall(qn("si")))
        logs.append(f"XML parse phase: {time.time() - phase_start:.2f}s.")

        scan_start = time.time()
        shared_indices: set[int] = set()
        shared_value_nodes_by_index: dict[int, list[tuple[str, ET.Element]]] = {}
        sheet_roots: dict[str, ET.Element] = {}
        inline_entries: list[tuple[str, ET.Element, str, str]] = []
        scanned_cells = 0
        candidate_cells = 0
        skipped_target_language = 0

        for sheet_name, sheet_path in zip(targets, target_paths):
            sheet_root = ET.fromstring(archive.read(sheet_path))
            sheet_roots[sheet_path] = sheet_root
            queued_before = len(shared_indices) + len(inline_entries)
            skipped_before = skipped_target_language
            for cell in sheet_root.iter(qn("c")):
                scanned_cells += 1
                cell_type = cell.get("t")
                if cell_type == "s":
                    value_node = cell.find(qn("v"))
                    if value_node is None or value_node.text is None:
                        continue
                    try:
                        index = int(value_node.text)
                    except ValueError:
                        continue
                    if index < 0 or index >= len(shared_items):
                        raise XlsxXmlFallbackRequired(f"shared string index out of range in {sheet_name}")
                    text_node = plain_shared_text_node(shared_items[index])
                    if text_node is None:
                        raise XlsxXmlFallbackRequired(f"rich text shared string is used in selected sheet '{sheet_name}'")
                    text = normalize_text(text_node.text or "")
                    if should_translate(text, "s"):
                        candidate_cells += 1
                        if is_likely_target_language(text, target_language):
                            skipped_target_language += 1
                            continue
                        shared_indices.add(index)
                        shared_text_by_index[index] = text
                        shared_value_nodes_by_index.setdefault(index, []).append((sheet_path, value_node))
                elif cell_type == "inlineStr":
                    text_node = plain_inline_text_node(cell)
                    if text_node is None:
                        raise XlsxXmlFallbackRequired(f"rich inline string is used in selected sheet '{sheet_name}'")
                    original_text = text_node.text or ""
                    text = normalize_text(original_text)
                    if should_translate(text, "s"):
                        candidate_cells += 1
                        if is_likely_target_language(text, target_language):
                            skipped_target_language += 1
                            continue
                        inline_entries.append((sheet_path, text_node, original_text, text))
            queued_after = len(shared_indices) + len(inline_entries)
            logs.append(
                f"Sheet '{sheet_name}': queued {queued_after - queued_before} XML text item(s); "
                f"skipped {skipped_target_language - skipped_before} likely already in {target_language}."
            )

        shared_entries = [(index, shared_text_by_index[index]) for index in sorted(shared_indices)]
        queued_texts = [text for _, text in shared_entries] + [text for _, _, _, text in inline_entries]
        if not queued_texts and not skipped_target_language:
            raise TranslationError("No translatable text cells found in the selected sheet(s).")

        unique_texts = list(dict.fromkeys(queued_texts))
        logs.append(f"XML scan phase: {time.time() - scan_start:.2f}s.")
        logs.append(f"Scanned {scanned_cells} XML cell(s).")
        logs.append(f"Found {candidate_cells} text cell(s) after base filters.")
        if skipped_target_language:
            logs.append(f"Skipped {skipped_target_language} cell(s) likely already in {target_language}.")
        logs.append(
            f"Deduplicated {len(queued_texts)} queued XML text item(s) to {len(unique_texts)} unique text(s); "
            f"saved {len(queued_texts) - len(unique_texts)} duplicate translation(s)."
        )

        translations: list[str] = []
        cache_hits = cache_misses = 0
        if unique_texts:
            translations, cache_hits, cache_misses = translate_texts_with_cache(
                texts=unique_texts,
                target_language=target_language,
                provider=provider,
                model=model,
                api_key=api_key,
                base_url=base_url,
                logs=logs,
                max_workers=max_workers,
                max_items=max_items,
            )
        logs.append(f"Translation cache: {cache_hits} hit(s), {cache_misses} miss(es).")

        translations_by_text = dict(zip(unique_texts, translations))
        patch_start = time.time()
        if shared_root is not None:
            for index, original_text in shared_entries:
                new_shared_item = copy.deepcopy(shared_items[index])
                text_node = plain_shared_text_node(new_shared_item)
                if text_node is None:
                    raise XlsxXmlFallbackRequired("shared string became unsupported during patch")
                text_node.text = preserve_wrapping(original_text, translations_by_text[original_text])
                new_index = len(shared_items)
                shared_items.append(new_shared_item)
                shared_root.append(new_shared_item)
                for _, value_node in shared_value_nodes_by_index[index]:
                    value_node.text = str(new_index)
            shared_root.set("uniqueCount", str(len(shared_items)))
        for _, text_node, original_text, normalized_text in inline_entries:
            text_node.text = preserve_wrapping(original_text, translations_by_text[normalized_text])

        replacements: dict[str, bytes] = {}
        if shared_root is not None and shared_entries:
            replacements["xl/sharedStrings.xml"] = xml_bytes(shared_root)
        changed_sheet_paths = {sheet_path for sheet_path, _, _, _ in inline_entries}
        for nodes in shared_value_nodes_by_index.values():
            changed_sheet_paths.update(sheet_path for sheet_path, _ in nodes)
        for sheet_path in changed_sheet_paths:
            sheet_root = sheet_roots[sheet_path]
            replacements[sheet_path] = xml_bytes(sheet_root)

        job_id = uuid.uuid4().hex
        job_dir = PROCESSED_DIR / job_id
        job_dir.mkdir(parents=True, exist_ok=True)
        output_path = job_dir / f"{Path(original_name).stem}_{slugify(target_language)}.xlsx"
        write_xlsx_copy(archive, output_path, replacements)
        logs.append(f"XML patch/write phase: {time.time() - patch_start:.2f}s.")
        if not unique_texts:
            logs.append("No cells were sent to the provider because all candidates looked already translated.")
        logs.append(f"Saved translated workbook: {output_path.name}")
        logs.append(f"✅ Total system execution time: {time.time() - system_start_time:.2f} seconds.")
        return job_id, output_path, logs


def qn(tag: str, namespace: str = XML_NS) -> str:
    return f"{{{namespace}}}{tag}"


def resolve_sheet_paths(workbook_root: ET.Element, rels_root: ET.Element) -> dict[str, str]:
    rel_targets = {
        rel.get("Id"): rel.get("Target", "")
        for rel in rels_root.findall(f"{{{PACKAGE_REL_NS}}}Relationship")
        if rel.get("Id")
    }
    paths: dict[str, str] = {}
    sheets_node = workbook_root.find(qn("sheets"))
    if sheets_node is None:
        raise XlsxXmlFallbackRequired("workbook has no sheets node")
    for sheet in sheets_node.findall(qn("sheet")):
        name = sheet.get("name")
        rel_id = sheet.get(qn("id", REL_NS))
        target = rel_targets.get(rel_id or "")
        if not name or not target:
            continue
        if target.startswith("/"):
            path = target.lstrip("/")
        else:
            path = posixpath.normpath(posixpath.join("xl", target))
        paths[name] = path
    if not paths:
        raise XlsxXmlFallbackRequired("could not resolve worksheet paths")
    return paths


def plain_shared_text_node(shared_item: ET.Element) -> ET.Element | None:
    direct_texts = shared_item.findall(qn("t"))
    runs = shared_item.findall(qn("r"))
    if len(direct_texts) == 1 and not runs:
        return direct_texts[0]
    return None


def plain_inline_text_node(cell: ET.Element) -> ET.Element | None:
    inline_node = cell.find(qn("is"))
    if inline_node is None:
        return None
    direct_texts = inline_node.findall(qn("t"))
    runs = inline_node.findall(qn("r"))
    if len(direct_texts) == 1 and not runs:
        return direct_texts[0]
    return None


def xml_bytes(root: ET.Element) -> bytes:
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def write_xlsx_copy(source: zipfile.ZipFile, output_path: Path, replacements: dict[str, bytes]) -> None:
    with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as target:
        for info in source.infolist():
            data = replacements.get(info.filename)
            if data is None:
                data = source.read(info.filename)
            target.writestr(info, data)


def should_translate(value: object, data_type: str) -> bool:
    if not isinstance(value, str):
        return False
    text = value.strip()
    if not text or text.startswith("="):
        return False
    if data_type == "f":
        return False
    if len(text) < 2:
        return False
    if re.fullmatch(r"[\d\s.,:%()+\-_/\\]+", text):
        return False
    if re.fullmatch(r"https?://\S+|\S+@\S+\.\S+", text, flags=re.IGNORECASE):
        return False
    return any(char.isalpha() for char in text)


def save_workbook(workbook: Any, original_name: str, target_language: str) -> tuple[str, Path]:
    job_id = uuid.uuid4().hex
    job_dir = PROCESSED_DIR / job_id
    job_dir.mkdir(parents=True, exist_ok=True)
    output_name = f"{Path(original_name).stem}_{slugify(target_language)}.xlsx"
    output_path = job_dir / output_name
    workbook.save(output_path)
    return job_id, output_path


def is_likely_target_language(text: str, target_language: str) -> bool:
    stripped = text.strip()
    if len(stripped) < 4:
        return False

    script_re = SCRIPT_RANGES.get(target_language)
    if script_re and script_re.search(stripped):
        return True

    if target_language == "Vietnamese" and VIETNAMESE_DIACRITIC_RE.search(stripped):
        return True

    lower_text = strip_accents(stripped.lower())
    words = re.findall(r"[a-zA-Z]+", lower_text)
    if len(words) < 2:
        return False

    markers = TARGET_LANGUAGE_MARKERS.get(target_language)
    if not markers:
        return False

    marker_hits = sum(1 for word in words if word in markers)
    return marker_hits >= 2


def strip_accents(value: str) -> str:
    value = value.replace("đ", "d").replace("Đ", "D")
    normalized = unicodedata.normalize("NFD", value)
    return "".join(char for char in normalized if unicodedata.category(char) != "Mn")


def normalize_text(value: str) -> str:
    return value.replace("\r\n", "\n")


def preserve_wrapping(original: str, translated: str) -> str:
    if "\r\n" in original:
        return translated.replace("\n", "\r\n")
    return translated


def translate_texts_with_cache(
    *,
    texts: list[str],
    target_language: str,
    provider: str,
    model: str,
    api_key: str,
    base_url: str,
    logs: list[str],
    max_workers: int = DEFAULT_MAX_WORKERS,
    max_items: int = DEFAULT_MAX_ITEMS,
) -> tuple[list[str], int, int]:
    cached = get_cached_translations(
        texts=texts,
        provider=provider,
        base_url=base_url,
        model=model,
        target_language=target_language,
    )
    missing = [text for text in texts if text not in cached]
    if missing:
        translated_missing = translate_texts(
            texts=missing,
            target_language=target_language,
            provider=provider,
            model=model,
            api_key=api_key,
            base_url=base_url,
            logs=logs,
            max_workers=max_workers,
            max_items=max_items,
            max_chars=DEFAULT_MAX_CHARS,
        )
        save_cached_translations(
            translations=dict(zip(missing, translated_missing)),
            provider=provider,
            base_url=base_url,
            model=model,
            target_language=target_language,
        )
        cached.update(zip(missing, translated_missing))
    return [cached[text] for text in texts], len(texts) - len(missing), len(missing)


def cache_db_path() -> Path:
    return STORAGE_DIR / "translation_cache.sqlite"


def cache_key(
    *,
    provider: str,
    base_url: str,
    model: str,
    target_language: str,
    text: str,
) -> str:
    material = "\n".join(
        [
            provider,
            hash_text(base_url.strip().rstrip("/")),
            model,
            target_language,
            hash_text(text),
        ]
    )
    return hash_text(material)


def hash_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def ensure_cache_schema(connection: sqlite3.Connection) -> None:
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS translation_cache (
            cache_key TEXT PRIMARY KEY,
            provider TEXT NOT NULL,
            base_url_hash TEXT NOT NULL,
            model TEXT NOT NULL,
            target_language TEXT NOT NULL,
            source_hash TEXT NOT NULL,
            source_text TEXT NOT NULL,
            translated_text TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
        """
    )


def get_cached_translations(
    *,
    texts: list[str],
    provider: str,
    base_url: str,
    model: str,
    target_language: str,
) -> dict[str, str]:
    if not texts:
        return {}
    keys_by_text = {
        text: cache_key(
            provider=provider,
            base_url=base_url,
            model=model,
            target_language=target_language,
            text=text,
        )
        for text in texts
    }
    connection = sqlite3.connect(cache_db_path())
    try:
        ensure_cache_schema(connection)
        rows = connection.execute(
            f"SELECT cache_key, translated_text FROM translation_cache WHERE cache_key IN ({','.join('?' for _ in keys_by_text)})",
            list(keys_by_text.values()),
        ).fetchall()
    finally:
        connection.close()
    translations_by_key = dict(rows)
    return {
        text: translations_by_key[key]
        for text, key in keys_by_text.items()
        if key in translations_by_key
    }


def save_cached_translations(
    *,
    translations: dict[str, str],
    provider: str,
    base_url: str,
    model: str,
    target_language: str,
) -> None:
    if not translations:
        return
    base_url_hash = hash_text(base_url.strip().rstrip("/"))
    rows = []
    for source_text, translated_text in translations.items():
        rows.append(
            (
                cache_key(
                    provider=provider,
                    base_url=base_url,
                    model=model,
                    target_language=target_language,
                    text=source_text,
                ),
                provider,
                base_url_hash,
                model,
                target_language,
                hash_text(source_text),
                source_text,
                translated_text,
            )
        )
    connection = sqlite3.connect(cache_db_path())
    try:
        ensure_cache_schema(connection)
        connection.executemany(
            """
            INSERT OR REPLACE INTO translation_cache (
                cache_key,
                provider,
                base_url_hash,
                model,
                target_language,
                source_hash,
                source_text,
                translated_text
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        connection.commit()
    finally:
        connection.close()


def translate_texts(
    *,
    texts: list[str],
    target_language: str,
    provider: str,
    model: str,
    api_key: str,
    base_url: str,
    logs: list[str],
    max_workers: int = DEFAULT_MAX_WORKERS,
    max_items: int = DEFAULT_MAX_ITEMS,
    max_chars: int = DEFAULT_MAX_CHARS,
) -> list[str]:
    translations: list[str] = []
    batches = list(batch_texts(texts, max_items=max_items, max_chars=max_chars))
    total_batches = len(batches)
    
    # Pre-allocate to maintain the original order since futures complete out of order
    results_by_index: list[list[str] | None] = [None] * total_batches
    
    import httpx

    timeout = httpx.Timeout(130.0, connect=30.0)
    limits = httpx.Limits(max_connections=max_workers + 2, max_keepalive_connections=max_workers + 2)
    with httpx.Client(timeout=timeout, limits=limits) as client, ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_index = {}
        for i, batch in enumerate(batches):
            batch_index = i + 1
            logs.append(
                f"Queuing batch {batch_index}/{total_batches}: "
                f"{len(batch)} text item(s), {sum(len(item) for item in batch)} chars."
            )
            future = executor.submit(
                translate_batch_with_split_retry,
                batch=batch,
                target_language=target_language,
                provider=provider,
                model=model,
                api_key=api_key,
                base_url=base_url,
                logs=logs,
                client=client,
            )
            future_to_index[future] = (i, batch_index, batch)
            
        for future in as_completed(future_to_index):
            i, batch_index, batch = future_to_index[future]
            try:
                result = future.result()
            except Exception as exc:
                if isinstance(exc, TranslationError):
                    raise
                raise TranslationError(f"Error in batch {batch_index}: {exc}") from exc
                
            if len(result) != len(batch):
                if len(batch) == 1 and len(result) > 1:
                    logs.append(
                        f"Provider returned {len(result)} fragments for one cell in batch {batch_index}; merging them."
                    )
                    result = ["\n".join(part.strip() for part in result if part.strip())]
                else:
                    raise TranslationError(
                        f"Translation count mismatch in batch {batch_index}: expected {len(batch)}, got {len(result)}."
                    )
            if len(result) != len(batch):
                raise TranslationError(
                    f"Translation count mismatch in batch {batch_index}: expected {len(batch)}, got {len(result)}."
                )
            results_by_index[i] = result
            logs.append(f"Completed batch {batch_index}/{total_batches}.")

    for result in results_by_index:
        if result is not None:
            translations.extend(result)
            
    logs.append(f"Translated {len(translations)} cell(s).")
    return translations


def translate_batch_with_split_retry(
    *,
    batch: list[str],
    target_language: str,
    provider: str,
    model: str,
    api_key: str,
    base_url: str,
    logs: list[str],
    client: Any,
) -> list[str]:
    for attempt in range(3):
        try:
            req_start_time = time.time()
            if provider == "gemini":
                result = call_gemini(batch, target_language, model, api_key, client)
            else:
                result = call_openai_chat(batch, target_language, provider, model, api_key, base_url, logs, client)
            
            req_elapsed = time.time() - req_start_time
            logs.append(f"Provider responded in {req_elapsed:.2f}s for {len(batch)} cell(s).")
            return result
        except ProviderRequestError as exc:
            if exc.status_code == 429 and attempt < 2:
                logs.append(f"Rate limit (429) on {len(batch)} cells. Sleeping 25s before retry {attempt+1}/3...")
                time.sleep(25)
                continue
            raise
        except TranslationError as exc:
            if not is_timeout_error(exc) or len(batch) == 1:
                raise
            break

    midpoint = len(batch) // 2
    logs.append(f"Provider timed out on {len(batch)} cell(s). Retrying as {midpoint} + {len(batch) - midpoint}.")
    first = translate_batch_with_split_retry(
        batch=batch[:midpoint],
        target_language=target_language,
        provider=provider,
        model=model,
        api_key=api_key,
        base_url=base_url,
        logs=logs,
        client=client,
    )
    second = translate_batch_with_split_retry(
        batch=batch[midpoint:],
        target_language=target_language,
        provider=provider,
        model=model,
        api_key=api_key,
        base_url=base_url,
        logs=logs,
        client=client,
    )
    return first + second


def is_timeout_error(exc: Exception) -> bool:
    message = str(exc).lower()
    return "timeout" in message or "took too long" in message or "timed out" in message


def batch_texts(texts: list[str], *, max_items: int = DEFAULT_MAX_ITEMS, max_chars: int = DEFAULT_MAX_CHARS) -> list[list[str]]:
    batches: list[list[str]] = []
    current: list[str] = []
    current_chars = 0
    for text in texts:
        item_chars = len(text)
        if current and (len(current) >= max_items or current_chars + item_chars > max_chars):
            batches.append(current)
            current = []
            current_chars = 0
        current.append(text)
        current_chars += item_chars
    if current:
        batches.append(current)
    return batches


def call_openai_chat(
    texts: list[str],
    target_language: str,
    provider: str,
    model: str,
    api_key: str,
    base_url: str,
    logs: list[str],
    client: Any,
) -> list[str]:
    chat_url = "https://api.openai.com/v1/chat/completions"
    models_url = "https://api.openai.com/v1/models"
    if provider == "aishop24h":
        chat_url = "https://aishop24h.com/v1/chat/completions"
        models_url = "https://aishop24h.com/v1/models"
        if base_url.strip():
            chat_url = f"{base_url.rstrip('/')}/chat/completions"
            models_url = f"{base_url.rstrip('/')}/models"
    elif provider == "openai_compatible":
        if not base_url.strip():
            raise TranslationError("Base URL is required for OpenAI-compatible providers.")
        chat_url = f"{base_url.rstrip('/')}/chat/completions"
        models_url = f"{base_url.rstrip('/')}/models"

    payload = {
        "model": model,
        "temperature": 0,
        "response_format": {"type": "json_object"},
        "messages": [
            {
                "role": "system",
                "content": (
                    "You translate spreadsheet cell text. Preserve meaning, numbers, placeholders, URLs, "
                    "line breaks, punctuation style, and do not add explanations."
                ),
            },
            {"role": "user", "content": build_prompt(texts, target_language)},
        ],
    }
    try:
        data = post_json(chat_url, payload, api_key, client)
    except ProviderRequestError as exc:
        if provider != "aishop24h" or exc.provider_code not in {"model_not_available", "model_not_found"}:
            raise
        available_models = list_openai_compatible_models(models_url, api_key, client)
        fallback_model = pick_fallback_model(available_models)
        if not fallback_model:
            raise TranslationError(
                f"{exc}. Could not find a fallback model from the provider's /models response."
            ) from exc
        logs.append(f"Model '{model}' is not available. Retrying with '{fallback_model}'.")
        payload["model"] = fallback_model
        data = post_json(chat_url, payload, api_key, client)
    content = extract_completion_text(data)
    return parse_translations(content)


def call_gemini(texts: list[str], target_language: str, model: str, api_key: str, client: Any) -> list[str]:
    encoded_model = urllib.parse.quote(model, safe="")
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{encoded_model}:generateContent?key={api_key}"
    payload = {
        "contents": [
            {
                "role": "user",
                "parts": [
                    {
                        "text": (
                            "Translate spreadsheet cell text. Return JSON only.\n\n"
                            + build_prompt(texts, target_language)
                        )
                    }
                ],
            }
        ],
        "generationConfig": {"temperature": 0, "responseMimeType": "application/json"},
    }
    data = post_json(url, payload, None, client)
    content = data["candidates"][0]["content"]["parts"][0]["text"]
    return parse_translations(content)


def build_prompt(texts: list[str], target_language: str) -> str:
    return json.dumps(
        {
            "instruction": (
                f"Translate each input item to {target_language}. Return exactly one output string per input item. "
                "Do not split one input item into multiple translations. Keep the same array length and order."
            ),
            "output_schema": {"translations": ["translated string"]},
            "expected_count": len(texts),
            "texts": texts,
        },
        ensure_ascii=False,
    )


def post_json(url: str, payload: dict[str, Any], api_key: str | None, client: Any) -> dict[str, Any]:
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "User-Agent": "QuickMediaTools/1.0 (+https://localhost)",
    }
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    try:
        response = client.post(url, headers=headers, json=payload)
        if response.status_code >= 400:
            raise ProviderRequestError(response.status_code, response.text)
        return response.json()
    except ProviderRequestError:
        raise
    except Exception as exc:
        raise TranslationError(f"Provider request failed: {exc}") from exc


def get_json(url: str, api_key: str | None, client: Any) -> dict[str, Any]:
    headers = {
        "Accept": "application/json",
        "User-Agent": "QuickMediaTools/1.0 (+https://localhost)",
    }
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    try:
        response = client.get(url, headers=headers)
        if response.status_code >= 400:
            raise ProviderRequestError(response.status_code, response.text)
        return response.json()
    except ProviderRequestError:
        raise
    except Exception as exc:
        raise TranslationError(f"Provider request failed: {exc}") from exc


def list_openai_compatible_models(url: str, api_key: str, client: Any) -> list[str]:
    data = get_json(url, api_key, client)
    models = data.get("data", [])
    ids = []
    for model in models:
        if isinstance(model, dict) and isinstance(model.get("id"), str):
            ids.append(model["id"])
    return ids


def pick_fallback_model(models: list[str]) -> str | None:
    preferred_exact = [
        "google/gemini-3-pro-preview",
        "google/gemini-3-flash-preview",
        "google/gemini-2.5-pro",
        "anthropic/claude-sonnet-4.6",
        "anthropic/claude-sonnet-4.5",
        "anthropic/claude-haiku-4.5",
    ]
    for model in preferred_exact:
        if model in models:
            return model

    preferred_terms = ("gemini", "claude", "gpt")
    for term in preferred_terms:
        for model in models:
            if term in model.lower() and "image" not in model.lower():
                return model
    return models[0] if models else None


def extract_provider_error_code(detail: str) -> str:
    try:
        data = json.loads(detail)
    except json.JSONDecodeError:
        return ""
    error = data.get("error")
    if isinstance(error, dict) and isinstance(error.get("code"), str):
        return error["code"]
    return ""


def parse_translations(content: str) -> list[str]:
    cleaned = content.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned)
        cleaned = re.sub(r"\s*```$", "", cleaned)
    data = json.loads(cleaned)
    translations = data.get("translations")
    if not isinstance(translations, list) or not all(isinstance(item, str) for item in translations):
        raise TranslationError("Provider returned invalid translation JSON.")
    return translations


def extract_completion_text(data: dict[str, Any]) -> str:
    choices = data.get("choices")
    if isinstance(choices, list) and choices:
        first = choices[0]
        if isinstance(first, dict):
            message = first.get("message")
            if isinstance(message, dict):
                content = message.get("content")
                if isinstance(content, str):
                    return content
            text = first.get("text")
            if isinstance(text, str):
                return text

    candidates = data.get("candidates")
    if isinstance(candidates, list) and candidates:
        first = candidates[0]
        if isinstance(first, dict):
            content = first.get("content")
            if isinstance(content, dict):
                parts = content.get("parts")
                if isinstance(parts, list):
                    text = "".join(
                        part.get("text", "")
                        for part in parts
                        if isinstance(part, dict) and isinstance(part.get("text"), str)
                    )
                    if text:
                        return text

    error = data.get("error")
    if isinstance(error, dict):
        message = error.get("message") or json.dumps(error, ensure_ascii=False)
        raise TranslationError(f"Provider returned an error response: {message}")

    raise TranslationError(
        "Provider response did not contain translated text. "
        f"Response keys: {', '.join(data.keys()) or 'none'}; sample: {json.dumps(data, ensure_ascii=False)[:800]}"
    )


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return slug or "translated"
